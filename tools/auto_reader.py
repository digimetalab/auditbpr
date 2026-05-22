#!/usr/bin/env python3
"""
BPR Data Auto-Reader — BPR Audit Intelligence System
=====================================================
Automatically detects and parses BPR financial data files.
Supports Excel (.xlsx, .xls) and CSV (.csv) formats.

Detection is based on CONTENT keywords, not filenames.
Output: Normalized JSON files per detected type.

Usage:
  python tools/auto_reader.py --data-dir ./data --output ./output/parsed
  python tools/auto_reader.py --validate-only --data-dir ./data
  python tools/auto_reader.py --demo
"""

import json
import sys
import os
import csv
import argparse
import datetime
import logging
import tempfile
import shutil
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

# ───────────────────────────────────────────────
# OPTIONAL DEPENDENCY IMPORTS
# ───────────────────────────────────────────────
_HAS_PANDAS = False
_HAS_OPENPYXL = False

try:
    import pandas as pd
    _HAS_PANDAS = True
except ImportError:
    pass

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    pass


# ───────────────────────────────────────────────
# LOGGING SETUP
# ───────────────────────────────────────────────
def setup_logging(output_dir: str) -> logging.Logger:
    """Configure logging to file and console."""
    logger = logging.getLogger("auto_reader")
    logger.setLevel(logging.DEBUG)

    # Clear any existing handlers to avoid duplicates
    logger.handlers.clear()

    # Console handler
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter("%(levelname)s: %(message)s"))
    logger.addHandler(ch)

    # File handler
    os.makedirs(output_dir, exist_ok=True)
    fh = logging.FileHandler(
        os.path.join(output_dir, "parsing_log.txt"),
        encoding="utf-8",
        mode="a",
    )
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter(
        "%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    ))
    logger.addHandler(fh)

    return logger


# ───────────────────────────────────────────────
# FILE TYPE DETECTION KEYWORDS
# ───────────────────────────────────────────────
FILE_TYPE_KEYWORDS: Dict[str, List[str]] = {
    "neraca": ["total aset", "total aktiva", "kewajiban", "ekuitas", "pasiva"],
    "laba_rugi": ["pendapatan bunga", "beban bunga", "laba rugi", "laba bersih"],
    "aset_produktif": ["kolektibilitas", "lancar", "kurang lancar", "diragukan", "macet"],
    "rasio": ["CAR", "KPMM", "NPL", "ROA", "BOPO", "LDR", "NIM"],
    "pengurus": ["direktur", "komisaris", "jabatan"],
    "pemegang_saham": ["pemegang saham", "kepemilikan", "saham"],
    "kap": ["kantor akuntan", "kap", "akuntan publik"],
}

# Minimum keyword matches required for each type
MIN_KEYWORD_MATCHES: Dict[str, int] = {
    "neraca": 2,
    "laba_rugi": 2,
    "aset_produktif": 3,
    "rasio": 3,
    "pengurus": 2,
    "pemegang_saham": 2,
    "kap": 1,
}


# ───────────────────────────────────────────────
# CONTENT EXTRACTION
# ───────────────────────────────────────────────
def extract_text_from_csv(filepath: str) -> str:
    """Extract all text content from a CSV file for keyword detection."""
    try:
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            return f.read().lower()
    except Exception:
        try:
            with open(filepath, "r", encoding="latin-1") as f:
                return f.read().lower()
        except Exception:
            return ""


def extract_text_from_excel(filepath: str) -> str:
    """Extract all text content from an Excel file for keyword detection."""
    if not _HAS_OPENPYXL:
        return ""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        text_parts: List[str] = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        text_parts.append(str(cell))
        wb.close()
        return " ".join(text_parts).lower()
    except Exception:
        return ""


def extract_text(filepath: str) -> str:
    """Extract text from any supported file type."""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".csv":
        return extract_text_from_csv(filepath)
    elif ext in (".xlsx", ".xls"):
        return extract_text_from_excel(filepath)
    return ""


# ───────────────────────────────────────────────
# FILE TYPE DETECTION
# ───────────────────────────────────────────────
def detect_file_type(filepath: str) -> Tuple[Optional[str], Dict[str, int]]:
    """
    Detect the type of a BPR data file based on content keywords.

    Returns:
        Tuple of (detected_type or None, match_counts_per_type)
    """
    content = extract_text(filepath)
    if not content:
        return None, {}

    match_counts: Dict[str, int] = {}
    for file_type, keywords in FILE_TYPE_KEYWORDS.items():
        count = 0
        for keyword in keywords:
            if keyword.lower() in content:
                count += 1
        match_counts[file_type] = count

    # Find the type with the highest match count
    best_type: Optional[str] = None
    best_count = 0
    for file_type, count in match_counts.items():
        min_required = MIN_KEYWORD_MATCHES.get(file_type, 2)
        if count >= min_required and count > best_count:
            best_type = file_type
            best_count = count

    return best_type, match_counts


# ───────────────────────────────────────────────
# CSV PARSING
# ───────────────────────────────────────────────
def parse_csv_file(filepath: str) -> List[Dict[str, Any]]:
    """
    Parse a CSV file into a list of row dicts.

    Handles common BPR data formats:
    - Comma and semicolon delimiters
    - Indonesian number formats (1.234.567 or 1,234,567)
    - UTF-8 and Latin-1 encodings
    """
    rows: List[Dict[str, Any]] = []

    # Try to read with different encodings
    content = ""
    for encoding in ["utf-8", "latin-1", "cp1252"]:
        try:
            with open(filepath, "r", encoding=encoding) as f:
                content = f.read()
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    if not content.strip():
        return rows

    # Detect delimiter
    first_lines = content.split("\n")[:5]
    semicolons = sum(line.count(";") for line in first_lines)
    commas = sum(line.count(",") for line in first_lines)
    delimiter = ";" if semicolons > commas else ","

    reader = csv.DictReader(content.strip().splitlines(), delimiter=delimiter)
    for row in reader:
        clean_row: Dict[str, Any] = {}
        for key, value in row.items():
            if key is None:
                continue
            clean_key = key.strip()
            if value is None:
                clean_row[clean_key] = None
            else:
                clean_row[clean_key] = _parse_value(value.strip())
        rows.append(clean_row)

    return rows


def _parse_value(value: str) -> Any:
    """Parse a string value, converting numbers where possible."""
    if not value or value == "-":
        return None

    # Try parsing as number (handle Indonesian format: 1.234.567,89)
    cleaned = value.replace(" ", "")

    # Check if it looks like a percentage
    if cleaned.endswith("%"):
        try:
            return float(cleaned[:-1].replace(",", "."))
        except ValueError:
            return value

    # Indonesian number format: dots as thousands, comma as decimal
    if "." in cleaned and "," in cleaned:
        try:
            return float(cleaned.replace(".", "").replace(",", "."))
        except ValueError:
            pass

    # Standard number format
    try:
        if "." in cleaned:
            return float(cleaned)
        return int(cleaned)
    except ValueError:
        pass

    return value


# ───────────────────────────────────────────────
# EXCEL PARSING
# ───────────────────────────────────────────────
def parse_excel_file(filepath: str) -> List[Dict[str, Any]]:
    """
    Parse an Excel file into a list of row dicts.

    Requires pandas and openpyxl.
    Falls back to openpyxl-only parsing if pandas is unavailable.
    """
    if _HAS_PANDAS:
        return _parse_excel_pandas(filepath)
    elif _HAS_OPENPYXL:
        return _parse_excel_openpyxl(filepath)
    else:
        return []


def _parse_excel_pandas(filepath: str) -> List[Dict[str, Any]]:
    """Parse Excel using pandas."""
    try:
        # Read first sheet by default
        df = pd.read_excel(filepath, engine="openpyxl")
        # Drop fully empty rows and columns
        df = df.dropna(how="all").dropna(axis=1, how="all")
        # Convert to list of dicts
        records = df.to_dict(orient="records")
        # Clean up NaN values
        clean_records: List[Dict[str, Any]] = []
        for record in records:
            clean: Dict[str, Any] = {}
            for k, v in record.items():
                key = str(k).strip()
                if _HAS_PANDAS and pd.isna(v):
                    clean[key] = None
                else:
                    clean[key] = v
            clean_records.append(clean)
        return clean_records
    except Exception:
        return []


def _parse_excel_openpyxl(filepath: str) -> List[Dict[str, Any]]:
    """Parse Excel using openpyxl only (fallback)."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return []

        rows_data: List[List[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows_data.append(list(row))
        wb.close()

        if len(rows_data) < 2:
            return []

        # First non-empty row as headers
        headers = [str(h).strip() if h is not None else f"col_{i}"
                   for i, h in enumerate(rows_data[0])]

        records: List[Dict[str, Any]] = []
        for row in rows_data[1:]:
            if all(v is None for v in row):
                continue
            record: Dict[str, Any] = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    record[headers[i]] = val
            records.append(record)

        return records
    except Exception:
        return []


# ───────────────────────────────────────────────
# DATA NORMALIZATION
# ───────────────────────────────────────────────
def normalize_data(rows: List[Dict[str, Any]], file_type: str) -> Dict[str, Any]:
    """
    Normalize parsed rows into a structured JSON object.

    Parameters:
        rows: Raw parsed rows from CSV/Excel
        file_type: Detected file type (neraca, laba_rugi, etc.)

    Returns:
        Normalized data dict with metadata and rows
    """
    return {
        "type": file_type,
        "type_label": _type_label(file_type),
        "row_count": len(rows),
        "columns": list(rows[0].keys()) if rows else [],
        "data": rows,
        "parsed_at": datetime.datetime.now().isoformat(),
    }


def _type_label(file_type: str) -> str:
    """Get human-readable label for a file type."""
    labels = {
        "neraca": "Balance Sheet (Neraca)",
        "laba_rugi": "Profit & Loss (Laba Rugi)",
        "aset_produktif": "Productive Asset Quality (Kualitas Aset Produktif)",
        "rasio": "Financial Ratios (Rasio Keuangan)",
        "pengurus": "Management (Pengurus)",
        "pemegang_saham": "Shareholders (Pemegang Saham)",
        "kap": "External Auditor (KAP)",
    }
    return labels.get(file_type, file_type)


# ───────────────────────────────────────────────
# DIRECTORY PROCESSING
# ───────────────────────────────────────────────
def process_directory(
    data_dir: str,
    output_dir: str,
    validate_only: bool = False,
    logger: Optional[logging.Logger] = None,
) -> Dict[str, Any]:
    """
    Process all supported files in a data directory.

    Parameters:
        data_dir: Path to directory containing data files
        output_dir: Path to output directory for parsed JSON
        validate_only: If True, only detect types without parsing
        logger: Logger instance

    Returns:
        Summary dict with parse results per file
    """
    if logger is None:
        logger = logging.getLogger("auto_reader")

    supported_extensions = {".xlsx", ".xls", ".csv"}
    results: List[Dict[str, Any]] = []

    if not os.path.isdir(data_dir):
        logger.error(f"Data directory not found: {data_dir}")
        return {"error": f"Data directory not found: {data_dir}", "files": []}

    # Find all supported files
    data_files: List[str] = []
    for entry in sorted(os.listdir(data_dir)):
        ext = os.path.splitext(entry)[1].lower()
        if ext in supported_extensions:
            data_files.append(os.path.join(data_dir, entry))

    if not data_files:
        logger.warning(f"No supported files found in {data_dir}")
        return {"warning": "No supported files found", "files": []}

    logger.info(f"Found {len(data_files)} file(s) in {data_dir}")

    # Check for Excel dependencies
    excel_files = [f for f in data_files
                   if os.path.splitext(f)[1].lower() in (".xlsx", ".xls")]
    if excel_files and not _HAS_OPENPYXL:
        logger.warning(
            "Excel files found but openpyxl is not installed. "
            "Install with: pip install openpyxl"
        )
    if excel_files and not _HAS_PANDAS:
        logger.warning(
            "pandas not installed — using openpyxl-only fallback for Excel. "
            "Install with: pip install pandas"
        )

    os.makedirs(output_dir, exist_ok=True)

    for filepath in data_files:
        filename = os.path.basename(filepath)
        logger.info(f"Processing: {filename}")

        # Detect file type
        file_type, match_counts = detect_file_type(filepath)

        file_result: Dict[str, Any] = {
            "filename": filename,
            "filepath": filepath,
            "detected_type": file_type,
            "match_counts": match_counts,
            "status": "unknown",
        }

        if file_type is None:
            file_result["status"] = "undetected"
            logger.warning(f"  Could not detect type for: {filename}")
            logger.debug(f"  Match counts: {match_counts}")
            results.append(file_result)
            continue

        logger.info(f"  Detected type: {file_type} ({_type_label(file_type)})")

        if validate_only:
            file_result["status"] = "validated"
            results.append(file_result)
            continue

        # Parse the file
        try:
            ext = os.path.splitext(filepath)[1].lower()
            if ext == ".csv":
                rows = parse_csv_file(filepath)
            elif ext in (".xlsx", ".xls"):
                rows = parse_excel_file(filepath)
            else:
                rows = []

            if not rows:
                file_result["status"] = "empty"
                logger.warning(f"  No data rows parsed from: {filename}")
                results.append(file_result)
                continue

            # Normalize and save
            normalized = normalize_data(rows, file_type)
            output_filename = f"parsed_data_{file_type}.json"
            output_path = os.path.join(output_dir, output_filename)

            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(normalized, f, ensure_ascii=False, indent=2, default=str)

            file_result["status"] = "success"
            file_result["output_file"] = output_filename
            file_result["row_count"] = len(rows)
            file_result["columns"] = normalized["columns"]
            logger.info(f"  [OK] Parsed {len(rows)} rows -> {output_filename}")

        except Exception as e:
            file_result["status"] = "error"
            file_result["error"] = str(e)
            logger.error(f"  [FAIL] Error parsing {filename}: {e}")

        results.append(file_result)

    return {"files": results, "processed_at": datetime.datetime.now().isoformat()}


# ───────────────────────────────────────────────
# SUMMARY GENERATION
# ───────────────────────────────────────────────
def generate_summary(
    results: Dict[str, Any],
    output_dir: str,
    logger: Optional[logging.Logger] = None,
) -> str:
    """
    Generate _summary.json with parse status per file.

    Returns:
        Path to the generated summary file
    """
    if logger is None:
        logger = logging.getLogger("auto_reader")

    summary_path = os.path.join(output_dir, "_summary.json")
    os.makedirs(output_dir, exist_ok=True)

    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2, default=str)

    logger.info(f"Summary written to: {summary_path}")
    return summary_path


# ───────────────────────────────────────────────
# DEMO MODE
# ───────────────────────────────────────────────
def run_demo(output_dir: str = "./output/parsed") -> None:
    """Generate sample BPR data files and process them."""
    print("=" * 60)
    print("BPR Data Auto-Reader — Demo Mode")
    print("=" * 60)

    # Create temporary demo data directory
    demo_dir = os.path.join(output_dir, "_demo_data")
    os.makedirs(demo_dir, exist_ok=True)

    # Sample Balance Sheet (Neraca)
    neraca_csv = """Pos,2022,2023,2024
Kas,2500,2800,3200
Penempatan pada Bank Lain,15000,18000,20000
Kredit yang Diberikan,85000,95000,110000
Penyisihan Penghapusan,-2500,-3000,-3500
Kredit Neto,82500,92000,106500
Aset Tetap,5000,4800,4600
Aset Lainnya,3000,3400,3700
Total Aset,108000,121000,138000
Tabungan,25000,28000,32000
Deposito,45000,50000,55000
Total DPK,70000,78000,87000
Kewajiban Lainnya,8000,9000,10000
Total Kewajiban,78000,87000,97000
Modal Disetor,20000,20000,20000
Cadangan,3000,4000,5000
Laba Ditahan,7000,10000,16000
Total Ekuitas,30000,34000,41000"""

    # Sample P&L (Laba Rugi)
    laba_rugi_csv = """Pos,2022,2023,2024
Pendapatan Bunga,18000,20000,23000
Beban Bunga,7000,7500,8000
Pendapatan Bunga Bersih,11000,12500,15000
Provisi dan Komisi,1500,1800,2000
Pendapatan Operasional Lainnya,500,600,700
Beban CKPN,1200,1500,1800
Beban Administrasi dan Umum,5000,5500,6000
Beban Operasional Lainnya,2000,2200,2500
Laba Operasional,4800,5700,7400
Pendapatan Non-Operasional,200,250,300
Beban Non-Operasional,100,150,200
Laba Sebelum Pajak,4900,5800,7500
Pajak,1200,1450,1875
Laba Bersih,3700,4350,5625
Laba Rugi Tahun Berjalan,3700,4350,5625"""

    # Sample Ratios (Rasio)
    rasio_csv = """Rasio,2022,2023,2024
CAR / KPMM (%),25.00,23.50,24.80
NPL Gross (%),3.20,4.10,3.50
NPL Net (%),2.10,2.80,2.30
ROA (%),3.40,3.60,4.10
BOPO (%),84.50,82.30,79.50
NIM (%),12.50,13.00,13.80
LDR (%),78.50,80.20,82.50
Cash Ratio (%),8.50,7.90,8.20"""

    # Write demo files
    for name, content in [
        ("demo_neraca.csv", neraca_csv),
        ("demo_laba_rugi.csv", laba_rugi_csv),
        ("demo_rasio.csv", rasio_csv),
    ]:
        filepath = os.path.join(demo_dir, name)
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"  + Created demo file: {name}")

    print()

    # Process demo data
    logger = setup_logging(output_dir)
    results = process_directory(demo_dir, output_dir, logger=logger)
    summary_path = generate_summary(results, output_dir, logger=logger)

    # Print results
    print("\n" + "-" * 60)
    print("RESULTS:")
    print("-" * 60)
    for file_info in results.get("files", []):
        status_icon = {
            "success": "[OK]",
            "validated": "[?]",
            "undetected": "[??]",
            "empty": "[!]",
            "error": "[FAIL]",
        }.get(file_info["status"], "?")
        print(f"  {status_icon} {file_info['filename']}")
        print(f"     Type   : {file_info.get('detected_type', 'N/A')}")
        print(f"     Status : {file_info['status']}")
        if file_info.get("row_count"):
            print(f"     Rows   : {file_info['row_count']}")

    print(f"\n  Summary: {summary_path}")

    # Cleanup demo data
    shutil.rmtree(demo_dir, ignore_errors=True)

    print("\n" + "=" * 60)
    print("Demo complete.")


# ───────────────────────────────────────────────
# CLI ENTRY POINT
# ───────────────────────────────────────────────
def main() -> None:
    """CLI entry point for auto_reader."""
    parser = argparse.ArgumentParser(
        description="BPR Data Auto-Reader — Parse BPR financial data files",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python tools/auto_reader.py --data-dir ./data --output ./output/parsed
  python tools/auto_reader.py --validate-only --data-dir ./data
  python tools/auto_reader.py --demo

Supported file types: .xlsx, .xls, .csv
Detection is based on content keywords, not filenames.

Dependencies:
  pip install pandas openpyxl    # For Excel support
  CSV parsing works without any external dependencies.
""",
    )
    parser.add_argument(
        "--data-dir",
        default="./data",
        help="Directory containing data files (default: ./data)",
    )
    parser.add_argument(
        "--output",
        default="./output/parsed",
        help="Output directory for parsed JSON (default: ./output/parsed)",
    )
    parser.add_argument(
        "--validate-only",
        action="store_true",
        help="Only detect file types without parsing",
    )
    parser.add_argument(
        "--demo",
        action="store_true",
        help="Run demo with sample data",
    )

    args = parser.parse_args()

    if args.demo:
        run_demo(args.output)
        return

    # Check dependencies for Excel files
    if not _HAS_OPENPYXL:
        print("[!]  openpyxl not installed -- Excel (.xlsx/.xls) files will be skipped.")
        print("   Install: pip install openpyxl")
    if not _HAS_PANDAS:
        print("[!]  pandas not installed -- using basic Excel parser.")
        print("   Install: pip install pandas")

    logger = setup_logging(args.output)
    logger.info("=" * 60)
    logger.info("BPR Data Auto-Reader — Starting")
    logger.info(f"Data directory: {args.data_dir}")
    logger.info(f"Output directory: {args.output}")
    logger.info(f"Validate only: {args.validate_only}")
    logger.info("=" * 60)

    results = process_directory(
        data_dir=args.data_dir,
        output_dir=args.output,
        validate_only=args.validate_only,
        logger=logger,
    )
    generate_summary(results, args.output, logger=logger)

    # Print summary to console
    success = sum(1 for f in results.get("files", []) if f["status"] == "success")
    total = len(results.get("files", []))
    print(f"\n{'='*60}")
    print(f"Processed {total} file(s): {success} success")
    print(f"Output: {args.output}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
